# -*- coding: utf-8 -*-
import os
import re
import io
import time
from datetime import datetime, timezone, timedelta

import pandas as pd
import requests
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# ===== 設定 =====
RELEASE_TAG = "news-latest"
ASSET_NAME = "yahoo_news.xlsx"
SHEET_NAMES = [
    "ホンダ",
    "トヨタ",
    "マツダ",
    "スバル",
    "ダイハツ",
    "スズキ",
    "三菱自動車",
    "日産",  # 任意で追加
]

# 既定は上のリスト。環境変数 NEWS_KEYWORDS に「ホンダ,トヨタ,…」と入れたら上書き可能
def get_keywords() -> list[str]:
    env = os.getenv("NEWS_KEYWORDS")
    if env:
        # カンマ区切り or 改行で分割
        parts = [p.strip() for p in re.split(r"[,\n]", env) if p.strip()]
        return parts or SHEET_NAMES
    return SHEET_NAMES


# ===== ユーティリティ =====
def jst_now():
    return datetime.now(timezone(timedelta(hours=9)))


def jst_str(fmt="%Y/%m/%d %H:%M"):
    return jst_now().strftime(fmt)


# ===== Chrome（headless） =====
def make_driver() -> webdriver.Chrome:
    opts = Options()
    chrome_path = os.getenv("CHROME_PATH")  # Actionsで注入
    if chrome_path:
        opts.binary_location = chrome_path
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1280,2000")
    # 長期運用時の出し分け対策：UA固定でも良いが、古すぎると要素出し分けが起きる場合あり
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)


# ===== 引用元のクリーンアップ =====
DATE_RE = re.compile(r"(?:\d{4}/\d{1,2}/\d{1,2}|\d{1,2}/\d{1,2})\s*\d{1,2}[:：]\d{2}")


def clean_source_text(text: str) -> str:
    if not text:
        return ""
    t = text
    t = re.sub(r"[（(][^）)]+[）)]", "", t)      # （）内を削除
    t = DATE_RE.sub("", t)                       # 日付+時刻パターンを削除
    t = re.sub(r"^\d+\s*", "", t)                # 先頭の通し番号（例: "2 Merkmal"）
    t = re.sub(r"\s{2,}", " ", t).strip()        # 余分な空白整理
    return t


# ===== Yahoo!ニュース検索（1ページ分） =====
def scrape_yahoo(keyword: str) -> pd.DataFrame:
    """
    指定キーワードでYahoo!ニュース（検索）から タイトル/URL/投稿日/引用元 を取得（1ページ）
    """
    driver = make_driver()
    url = (
        f"https://news.yahoo.co.jp/search?p={keyword}"
        f"&ei=utf-8&categories=domestic,world,business,it,science,life,local"
    )
    driver.get(url)
    time.sleep(5)  # 初期描画待ち

    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    # li のクラスは変動しやすいので正規表現で拾う
    items = soup.find_all("li", class_=re.compile("sc-1u4589e-0"))
    rows = []
    for li in items:
        try:
            title_tag = li.find("div", class_=re.compile("sc-3ls169-0"))
            link_tag = li.find("a", href=True)
            time_tag = li.find("time")

            title = title_tag.get_text(strip=True) if title_tag else ""
            url = link_tag["href"] if link_tag else ""
            date_str = time_tag.get_text(strip=True) if time_tag else ""

            # 投稿日：フォーマットできれば "9/26 16:19" に正規化
            pub_date = "取得不可"
            if date_str:
                ds = re.sub(r'\([月火水木金土日]\)', '', date_str).strip()
                try:
                    # 入力フォーマットが「2024/09/26 16:19」の場合
                    dt = datetime.strptime(ds, "%Y/%m/%d %H:%M")
                    pub_date = dt.strftime("%-m/%-d %H:%M")
                except ValueError:
                    try:
                        # 入力フォーマットが「9/26 16:19」の場合
                        dt = datetime.strptime(ds, "%m/%d %H:%M")
                        pub_date = dt.strftime("%-m/%-d %H:%M")
                    except Exception:
                        # どちらにも当てはまらない場合、元の文字列をそのまま使用
                        pub_date = ds

            # 引用元（媒体＋カテゴリなど）を抽出してクリーン
            source = ""
            for sel in [
                "div.sc-n3vj8g-0.yoLqH div.sc-110wjhy-8.bsEjY span",
                "div.sc-n3vj8g-0.yoLqH",
                "span",
                "div",
            ]:
                el = li.select_one(sel)
                if not el:
                    continue
                raw = el.get_text(" ", strip=True)
                txt = clean_source_text(raw)
                if txt and not txt.isdigit():
                    source = txt
                    break

            if title and url:
                rows.append(
                    {
                        "タイトル": title,
                        "URL": url,
                        "投稿日": pub_date,
                        "引用元": source or "Yahoo",
                        "取得日時": jst_str(),       # 追記運用のため取得時刻も保持
                        "検索キーワード": keyword,   # 念のため列としても持っておく
                    }
                )
        except Exception:
            continue

    return pd.DataFrame(rows, columns=["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード"])


# ===== Releaseから既存Excelを取得（全シート） =====
def download_existing_book(repo: str, tag: str, asset_name: str, token: str) -> dict[str, pd.DataFrame]:
    """
    Release(tag)の既存Excel全シートを読み出して {sheet_name: df} で返す。
    見つからなければ、指定シート名それぞれ空DFで返す。
    """
    # 初期値（指定の全シート分、空DF）
    empty_cols = ["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード"]
    dfs: dict[str, pd.DataFrame] = {sn: pd.DataFrame(columns=empty_cols) for sn in SHEET_NAMES}

    if not (repo and tag):
        print("⚠️ download_existing_book: repo/tag が未設定のためスキップ")
        return dfs

    base = "https://api.github.com"
    headers = {"Accept": "application/vnd.github+json"}
    # token は browser_download_url では不要だが、/releases 読み出しにはあってもOK
    if token:
        headers["Authorization"] = f"Bearer {token}"

    # 1) Release 情報取得
    url_rel = f"{base}/repos/{repo}/releases/tags/{tag}"
    r = requests.get(url_rel, headers=headers)
    print(f"🔎 GET {url_rel} -> {r.status_code}")
    if r.status_code != 200:
        print("⚠️ Releaseが見つからないか、取得に失敗。既存は空として続行します。")
        return dfs
    rel = r.json()

    # 2) 対象アセット探索
    asset = next((a for a in rel.get("assets", []) if a.get("name") == asset_name), None)
    if not asset:
        print(f"⚠️ Releaseに {asset_name} が存在しません。既存は空として続行します。")
        return dfs

    # 3) ダウンロードは browser_download_url を使用（認証不要で安定）
    dl_url = asset.get("browser_download_url")
    if not dl_url:
        print("⚠️ browser_download_url が見つかりません。既存は空として続行します。")
        return dfs

    dr = requests.get(dl_url)
    print(f"⬇️  Download {dl_url} -> {dr.status_code}, {len(dr.content)} bytes")
    if dr.status_code != 200:
        print("⚠️ 既存Excelのダウンロードに失敗。既存は空として続行します。")
        return dfs

    # 4) Excel 読み込み
    with io.BytesIO(dr.content) as bio:
        try:
            book = pd.read_excel(bio, sheet_name=None)
        except Exception as e:
            print(f"⚠️ 既存Excelの読み込みに失敗: {e}")
            return dfs

    # 5) シートごとに整形して返す
    for sn in SHEET_NAMES:
        if sn in book:
            df = book[sn]
            # 欠けている列があれば補完（将来の列追加にも耐性）
            for col in empty_cols:
                if col not in df.columns:
                    df[col] = ""
            dfs[sn] = df[empty_cols].copy()

    return dfs


# ===== Excel保存（体裁調整つき） =====
def save_book_with_format(dfs: dict[str, pd.DataFrame], path: str):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    # 既定で作られる最初のシートを削除
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(title=sheet_name)
        # ヘッダー
        headers = ["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード"]
        ws.append(headers)
        # データ
        if not df.empty:
            for row in df[headers].itertuples(index=False, name=None):
                ws.append(list(row))

        # オートフィルター
        max_col = ws.max_column
        max_row = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # ヘッダー太字 & 縦中央
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        # 列幅（軽調整）
        widths = {
            "A": 50,  # タイトル
            "B": 60,  # URL
            "C": 16,  # 投稿日
            "D": 24,  # 引用元
            "E": 16,  # 取得日時
            "F": 16,  # 検索キーワード
        }
        for col, wdt in widths.items():
            if ws.max_column >= ord(col) - 64:
                ws.column_dimensions[col].width = wdt

        # 1行目固定
        ws.freeze_panes = "A2"

    # 出力
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)


# ===== メイン =====
def main():
    # キーワードは環境変数NEWS_KEYWORDSで上書き可能（例: "ホンダ,トヨタ,..."）
    keywords = get_keywords()
    print(f"🔎 キーワード一覧: {', '.join(keywords)}")

    # 1) 既存ブック取得（固定Releaseから）
    token = os.getenv("GITHUB_TOKEN", "")
    repo = os.getenv("GITHUB_REPOSITORY", "")
    dfs_old = download_existing_book(repo, RELEASE_TAG, ASSET_NAME, token)

    # 2) 新規スクレイプ → シートごとにマージ（URLで重複排除、既存優先＝新規は末尾）
    dfs_merged: dict[str, pd.DataFrame] = {}
    for kw in keywords:
        df_old = dfs_old.get(kw, pd.DataFrame(columns=["タイトル", "URL", "投稿日", "引用元", "取得日時", "検索キーワード"]))
        df_new = scrape_yahoo(kw)

        df_all = pd.concat([df_old, df_new], ignore_index=True)
        if not df_all.empty:
            df_all = df_all.dropna(subset=["URL"]).drop_duplicates(subset=["URL"], keep="first")
            # 並べ替えはしない：既存の並びを維持し、新規は末尾に付く
        dfs_merged[kw] = df_all

        print(f"  - {kw}: 既存 {len(df_old)} 件 + 新規 {len(df_new)} 件 → 合計 {len(df_all)} 件")

    # 3) 保存（各シートに出力、ヘッダにフィルター／フリーズ等）
    os.makedirs("output", exist_ok=True)
    out_path = os.path.join("output", ASSET_NAME)
    save_book_with_format(dfs_merged, out_path)

    print(f"✅ Excel出力: {out_path}")
    # 固定DLリンク（実リポジトリ名が分かれば整形）
    if repo:
        owner_repo = repo
    else:
        owner_repo = "<OWNER>/<REPO>"
    print(f"🔗 固定DL: https://github.com/{owner_repo}/releases/download/{RELEASE_TAG}/{ASSET_NAME}")


if __name__ == "__main__":
    main()
